# Script to crawl websites for pricing information and matching this with current pricing lists
# Any question? Send question to Mats Lundell-Nygjelten with email lundell.nygjelten@gmail.com

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl
import re
import time

def comparePrices():
	# Login information to Elektroskandia
	loginUsername = "<Username>"
	loginPassword = "<Password>"

	# Set path, open with openpyxl
	pathToFile = "<nameOfSpreadsheet>.xlsx"
	wb = openpyxl.load_workbook(pathToFile)
	sheet = wb.worksheets[0]
	row_count = sheet.max_row

	# Start the webdriver with the browser you are using
	driver = webdriver.Chrome()
	#driver.maximize_window()

	# Login to Elektroskandia
	driver.get("https://webshop.elektroskandia.no/nor/#")
	elemUsername = driver.find_element_by_xpath("//input[@id='j_username_leftAside']").send_keys(loginUsername)
	elemPassword = driver.find_element_by_xpath("//input[@id='j_password_leftAside']").send_keys(loginPassword)
	elemLoginButton = driver.find_element_by_xpath("//div[@class='width100 sign-in mb20']//button[@type='submit'][contains(text(),'Logg inn')]").click()
	elemSaveSettings = driver.find_element_by_xpath("//button[@class='rx-btn rx-btn-white closeSessionDrawer right']").click()

	# Setup the spreadsheet with names for the new columns
	sheet.cell(row = 1, column = 4).value = 'Live Pris'
	sheet.cell(row = 1, column = 5).value = 'Feil i nettoprisliste?'
	sheet.cell(row = 1, column = 6).value = 'På lager?'


	# Create a loop to go through every item and extract information
	for item in range(40, 50):

		if sheet.cell(row = item, column = 1).value:
			# Get the item number we want to use for search
			cellVarenummer = sheet.cell(row = item, column = 1).value

			# Find item and enter the item's page
			elemSearchbar = driver.find_element_by_id("search").send_keys(cellVarenummer)
			elemSearchButton = driver.find_element_by_xpath("//div[@class='controls']//button[@type='button']").click()
			try:
				elemClickOnProduct = driver.find_element_by_xpath("/html[1]/body[1]/div[6]/div[2]/div[3]/div[4]/div[2]/div[4]/div[1]/div[1]/div[1]/div[1]/a[1]/img[1]").click()
			except Exception:
				pass
			time.sleep(1)

			# Extract useful information
			# Live sales unit pricing
			stringWithLiveSalesUnitPrice = driver.find_element_by_xpath("/html[1]/body[1]/div[6]/div[2]/div[6]/div[1]/div[5]/div[2]/div[1]/div[3]/div[1]/span[1]").text
			liveSalesUnitPrice = stripPrice(stringWithLiveSalesUnitPrice)

			# Push extracted information to spreadsheet
			sheet.cell(row = item, column = 4).value = liveSalesUnitPrice

			# Checks if there is a difference from the price in pricing list and the live price
			if float(sheet.cell(row = item, column = 3).value) - float(sheet.cell(row = item, column = 4).value) != 0:
				sheet.cell(row = item, column = 5).value = "FORSKJELLIG PRIS"	

			availability = driver.find_element_by_xpath("/html[1]/body[1]/div[6]/div[2]/div[6]/div[1]/div[5]/div[4]/div[2]/div[2]").text
			# Check if item in stock
			if 'Ikke' in availability:
				sheet.cell(row = item, column = 6).value = "Nei"
				

	# Saves and updates spreadsheet with extracted information
	wb.save(pathToFile)

# Function to strip down live prices
def stripPrice(price):
	strippedPrice = ''
	for word in price:
		if word.isdigit():
			strippedPrice += word
		elif word == ',':
			strippedPrice += '.'

	return strippedPrice

