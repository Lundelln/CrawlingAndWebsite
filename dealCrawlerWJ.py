# Script to crawl websites for pricing information and matching this with current pricing lists
# Any question? Send question to Mats Lundell-Nygjelten with email lundell.nygjelten@gmail.com

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl
import re

def comparePrices():
	# Login information to shop.wj.no
	loginUsername = "<Username>"
	loginPassword = "<Password>"

	# Set path, open with openpyxl
	pathToFile = "<SpreadsheetName>.xlsx"
	wb = openpyxl.load_workbook(pathToFile)
	sheet = wb.worksheets[0]
	row_count = sheet.max_row

	# Start the webdriver with the browser you are using
	driver = webdriver.Chrome()
	#driver.maximize_window()

	# Login to shop.wj.no
	driver.get("https://shop.wj.no/")
	elemUsername = driver.find_element_by_name("j_username").send_keys(loginUsername)
	elemPassword = driver.find_element_by_name("j_password").send_keys(loginPassword)
	elemLoginButton = driver.find_element_by_xpath("//button[@type='submit']").click()

	# Setup the spreadsheet with names for the new columns
	sheet.cell(row = 4, column = 9).value = 'Live Pris'
	sheet.cell(row = 4, column = 10).value = 'Live Antall i salgsenhet'
	sheet.cell(row = 4, column = 11).value = 'Live Pris salgsenhet'
	sheet.cell(row = 4, column = 12).value = 'Feil i nettoprisliste?'

	# Create a loop to go through every item and extract information
	for item in range(40, 50):

		if sheet.cell(row = item, column = 1).value:
			# Get the item number we want to use for search
			cellVarenummer = sheet.cell(row = item, column = 1).value

			# Find item and enter the item's page
			elemSearchbar = driver.find_element_by_id("search").send_keys(cellVarenummer)
			elemSearchButton = driver.find_element_by_xpath("//button[@class='btn btn-dark']").click()
			# Use function to see if element exists
			try:
				elemClickOnProduct = driver.find_element_by_xpath("/html[1]/body[1]/div[3]/div[2]/div[4]/div[3]/div[3]/div[3]/div[1]/a[1]/div[1]/img[1]").click()

				# Extract useful information
				# Live per unit pricing
				stringWithLivePerUnitPrice = driver.find_element_by_class_name("perUnitPrice").text
				livePerUnitPrice = stripPrice(stringWithLivePerUnitPrice)

				# Live stack number 
				stringWithLiveStackNumber = driver.find_element_by_name("packSize").text
				liveStackNumber = stripPrice(stringWithLiveStackNumber)

				# Live sales unit pricing
				stringWithLiveSalesUnitPrice = driver.find_element_by_id("packSizes-prize-target").text
				liveSalesUnitPrice = stripPrice(stringWithLiveSalesUnitPrice)

				# Push extracted information to spreadsheet
				sheet.cell(row = item, column = 9).value = livePerUnitPrice
				sheet.cell(row = item, column = 10).value = liveStackNumber
				sheet.cell(row = item, column = 11).value = liveSalesUnitPrice

				# Checks if there is a difference from the price in pricing list and the live price
				if float(sheet.cell(row = item, column = 5).value) - float(sheet.cell(row = item, column = 9).value) != 0 or float(sheet.cell(row = item, column = 7).value) - float(sheet.cell(row = item, column = 11).value) != 0:
					sheet.cell(row = item, column = 12).value = "FORSKJELLIG PRIS"	

			# If this exception is raised, the item is not in the web shop
			except Exception:
				sheet.cell(row = item, column = 8).value = "Finner ikke varen"

	# Saves and updates spreadsheet with extracted information
	return wb.save(pathToFile)

# Function to strip down live prices
def stripPrice(price):
	strippedPrice = ''
	for word in price:
		if word.isdigit():
			strippedPrice += word
		elif word == ',':
			strippedPrice += '.'

	return strippedPrice

